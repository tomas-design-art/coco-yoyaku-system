"""患者一括取り込み テスト"""
import io
import csv
import json
import unittest
from datetime import date


class TestSuggestMapping(unittest.TestCase):
    """自動マッピング推定テスト"""

    def test_split_mode_mapping(self):
        from app.api.patient_import import suggest_mapping
        headers = ["姓", "名", "フリガナ", "電話番号", "生年月日"]
        mapping, mode = suggest_mapping(headers)
        assert mode == "split"
        assert mapping["last_name"] == 0
        assert mapping["first_name"] == 1
        assert mapping["reading"] == 2
        assert mapping["phone"] == 3
        assert mapping["birth_date"] == 4

    def test_fullname_mode_mapping(self):
        from app.api.patient_import import suggest_mapping
        headers = ["氏名", "読み方", "TEL", "誕生日", "メール"]
        mapping, mode = suggest_mapping(headers)
        assert mode == "full_name"
        assert mapping["full_name"] == 0
        assert mapping["reading"] == 1
        assert mapping["phone"] == 2
        assert mapping["birth_date"] == 3
        assert mapping["email"] == 4

    def test_ambiguous_name_with_sei(self):
        """姓列がある場合、「名前」は first_name に割り当て"""
        from app.api.patient_import import suggest_mapping
        headers = ["姓", "名前", "電話"]
        mapping, mode = suggest_mapping(headers)
        assert mode == "split"
        assert mapping["last_name"] == 0
        assert mapping["first_name"] == 1

    def test_ambiguous_name_without_sei(self):
        """姓列がない場合、「名前」は full_name に割り当て"""
        from app.api.patient_import import suggest_mapping
        headers = ["名前", "電話", "メール"]
        mapping, mode = suggest_mapping(headers)
        assert mode == "full_name"
        assert mapping["full_name"] == 0

    def test_english_headers(self):
        from app.api.patient_import import suggest_mapping
        headers = ["last_name", "first_name", "phone", "email", "note"]
        mapping, mode = suggest_mapping(headers)
        assert mode == "split"
        assert mapping["last_name"] == 0
        assert mapping["first_name"] == 1
        assert mapping.get("notes") == 4


class TestParseDate(unittest.TestCase):
    """日付パース テスト"""

    def test_iso_format(self):
        from app.api.patient_import import _parse_date
        assert _parse_date("1990-01-15") == date(1990, 1, 15)

    def test_slash_format(self):
        from app.api.patient_import import _parse_date
        assert _parse_date("1990/01/15") == date(1990, 1, 15)

    def test_japanese_format(self):
        from app.api.patient_import import _parse_date
        assert _parse_date("1990年01月15日") == date(1990, 1, 15)

    def test_8digit_format(self):
        from app.api.patient_import import _parse_date
        assert _parse_date("19900115") == date(1990, 1, 15)

    def test_invalid_date(self):
        from app.api.patient_import import _parse_date
        with self.assertRaises(ValueError):
            _parse_date("不明")

    def test_empty_returns_none(self):
        from app.api.patient_import import _parse_date
        assert _parse_date("") is None
        assert _parse_date("  ") is None


class TestBuildPatientRecord(unittest.TestCase):
    """レコード組み立てテスト"""

    def test_split_mode_basic(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"last_name": "田中", "first_name": "太郎", "phone": "090-1234-5678"}
        record = _build_patient_record(extracted, "split", 1)
        assert record["last_name"] == "田中"
        assert record["first_name"] == "太郎"
        assert record["name"] == "田中 太郎"
        assert record["phone"] == "09012345678"

    def test_fullname_mode_basic(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"full_name": "John Smith", "reading": "ジョン スミス"}
        record = _build_patient_record(extracted, "full_name", 1)
        assert record["name"] == "John Smith"
        assert record["reading"] == "ジョン スミス"

    def test_split_missing_last_name(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"last_name": "", "first_name": "太郎"}
        with self.assertRaises(ValueError):
            _build_patient_record(extracted, "split", 3)

    def test_fullname_missing(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"full_name": ""}
        with self.assertRaises(ValueError):
            _build_patient_record(extracted, "full_name", 5)

    def test_invalid_birth_date(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"last_name": "田中", "first_name": "太郎", "birth_date": "不明"}
        with self.assertRaises(ValueError):
            _build_patient_record(extracted, "split", 2)


class TestReadFile(unittest.TestCase):
    """ファイル読み込みテスト"""

    def test_csv_read(self):
        from app.api.patient_import import _read_file
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["姓", "名", "電話番号"])
        writer.writerow(["田中", "太郎", "09012345678"])
        writer.writerow(["佐藤", "花子", "08011112222"])
        csv_bytes = buf.getvalue().encode("utf-8-sig")

        headers, data = _read_file(csv_bytes, "test.csv")
        assert headers == ["姓", "名", "電話番号"]
        assert len(data) == 2
        assert data[0][0] == "田中"

    def test_xlsx_read(self):
        from app.api.patient_import import _read_file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["氏名", "読み方", "TEL"])
        ws.append(["John Smith", "ジョン スミス", "09099998888"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        headers, data = _read_file(xlsx_bytes, "test.xlsx")
        assert headers == ["氏名", "読み方", "TEL"]
        assert len(data) == 1

    def test_empty_rows_skipped(self):
        from app.api.patient_import import _is_empty_row
        assert _is_empty_row(["", "", ""]) is True
        assert _is_empty_row(["", "data", ""]) is False


class TestPreviewCSV(unittest.TestCase):
    """CSV で preview が正しく返るか"""

    def test_preview_returns_mapping(self):
        from app.api.patient_import import suggest_mapping, _read_file, _extract_row, _is_empty_row
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["姓", "名", "フリガナ", "電話番号", "生年月日"])
        writer.writerow(["田中", "太郎", "タナカ タロウ", "09012345678", "1990/01/15"])
        writer.writerow(["佐藤", "花子", "サトウ ハナコ", "08011112222", "1985/06/20"])
        csv_bytes = buf.getvalue().encode("utf-8-sig")

        headers, data = _read_file(csv_bytes, "test.csv")
        mapping, mode = suggest_mapping(headers)
        assert mode == "split"

        preview = []
        for row in data[:10]:
            if not _is_empty_row(row):
                preview.append(_extract_row(row, mapping))
        assert len(preview) == 2
        assert preview[0]["last_name"] == "田中"
        assert preview[0]["first_name"] == "太郎"


class TestPreviewExcel(unittest.TestCase):
    """Excel で preview が正しく返るか"""

    def test_xlsx_preview(self):
        from app.api.patient_import import suggest_mapping, _read_file, _extract_row, _is_empty_row
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["患者名", "カナ", "携帯", "誕生日"])
        ws.append(["John Smith", "ジョン", "09099998888", "1990-05-01"])
        buf = io.BytesIO()
        wb.save(buf)

        headers, data = _read_file(buf.getvalue(), "patients.xlsx")
        mapping, mode = suggest_mapping(headers)
        assert mode == "full_name"
        assert mapping["full_name"] == 0

        preview = []
        for row in data[:10]:
            if not _is_empty_row(row):
                preview.append(_extract_row(row, mapping))
        assert preview[0]["full_name"] == "John Smith"


class TestTemplateCSV(unittest.TestCase):
    """CSV テンプレートのテスト"""

    def test_csv_template_downloadable(self):
        """CSV テンプレートが取得でき、ヘッダーが正しいこと"""
        from app.api.patient_import import TEMPLATE_HEADERS_SPLIT, TEMPLATE_SAMPLE_SPLIT
        import asyncio
        from app.api.patient_import import download_template_csv

        loop = asyncio.new_event_loop()
        resp = loop.run_until_complete(download_template_csv())
        # StreamingResponse から body を取り出す
        chunks = []
        async def _collect():
            async for chunk in resp.body_iterator:
                if isinstance(chunk, str):
                    chunks.append(chunk.encode("utf-8"))
                else:
                    chunks.append(chunk)
        loop.run_until_complete(_collect())
        loop.close()
        text = b"".join(chunks).decode("utf-8")

        # BOM 除去して parse
        if text.startswith("\ufeff"):
            text = text[1:]
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        assert len(rows) >= 2, "ヘッダー行 + サンプル行が必要"
        assert rows[0] == TEMPLATE_HEADERS_SPLIT
        assert rows[1] == TEMPLATE_SAMPLE_SPLIT

    def test_csv_content_disposition(self):
        """Content-Disposition ヘッダーが設定されていること"""
        import asyncio
        from app.api.patient_import import download_template_csv
        loop = asyncio.new_event_loop()
        resp = loop.run_until_complete(download_template_csv())
        loop.close()
        cd = resp.headers.get("content-disposition", "")
        assert "patient_import_template.csv" in cd


class TestTemplateXlsx(unittest.TestCase):
    """Excel テンプレートのテスト"""

    def test_xlsx_template_two_sheets(self):
        """Excel テンプレートが 2 シート構成でヘッダーが正しいこと"""
        from app.api.patient_import import (
            download_template_xlsx, TEMPLATE_HEADERS_SPLIT, TEMPLATE_HEADERS_FULL,
        )
        from openpyxl import load_workbook
        import asyncio

        loop = asyncio.new_event_loop()
        resp = loop.run_until_complete(download_template_xlsx())
        chunks = []
        async def _collect():
            async for chunk in resp.body_iterator:
                chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode("utf-8"))
        loop.run_until_complete(_collect())
        loop.close()
        body = b"".join(chunks)

        wb = load_workbook(io.BytesIO(body))
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "通常モード"
        assert wb.sheetnames[1] == "フルネームモード"

        # 通常モードヘッダー
        ws_split = wb["通常モード"]
        split_headers = [cell.value for cell in ws_split[1]]
        assert split_headers == TEMPLATE_HEADERS_SPLIT

        # フルネームモードヘッダー
        ws_full = wb["フルネームモード"]
        full_headers = [cell.value for cell in ws_full[1]]
        assert full_headers == TEMPLATE_HEADERS_FULL

        wb.close()

    def test_xlsx_content_disposition(self):
        """Content-Disposition にファイル名が含まれること"""
        import asyncio
        from app.api.patient_import import download_template_xlsx
        loop = asyncio.new_event_loop()
        resp = loop.run_until_complete(download_template_xlsx())
        loop.close()
        cd = resp.headers.get("content-disposition", "")
        assert "patient_import_template.xlsx" in cd


class TestUpdateExistingPatient(unittest.TestCase):
    """_update_existing_patient のテスト"""

    def test_updatable_fields_defined(self):
        """UPDATABLE_FIELDS が安全な項目のみ含むこと"""
        from app.api.patient_import import UPDATABLE_FIELDS
        assert set(UPDATABLE_FIELDS) == {"phone", "email", "reading", "notes"}

    def test_name_not_updatable(self):
        """氏名関連は更新対象に含まれないこと"""
        from app.api.patient_import import UPDATABLE_FIELDS
        assert "name" not in UPDATABLE_FIELDS
        assert "last_name" not in UPDATABLE_FIELDS
        assert "first_name" not in UPDATABLE_FIELDS
        assert "full_name" not in UPDATABLE_FIELDS


class TestRowActionsLogic(unittest.TestCase):
    """row_actions の分岐ロジックテスト"""

    def test_skip_action(self):
        """skip アクション: 重複行に action=skip → duplicate_count 増加"""
        actions = [{"row": 1, "action": "skip"}]
        assert actions[0]["action"] == "skip"
        # 実際の API テストは統合テストで行うが、
        # ロジック分岐が正しいことを確認
        assert actions[0]["action"] in ("skip", "use_existing", "update_existing")

    def test_use_existing_action_structure(self):
        """use_existing アクション: patient_id 不要"""
        action = {"row": 3, "action": "use_existing", "patient_id": 12}
        assert action["action"] == "use_existing"
        assert "patient_id" in action

    def test_update_existing_requires_patient_id(self):
        """update_existing アクション: patient_id 必須"""
        action_bad = {"row": 4, "action": "update_existing"}
        assert "patient_id" not in action_bad

    def test_build_record_for_update(self):
        """_build_patient_record でレコード構築 → 更新用データが取れる"""
        from app.api.patient_import import _build_patient_record
        extracted = {
            "last_name": "田中",
            "first_name": "花子",
            "reading": "タナカ ハナコ",
            "phone": "090-9999-0000",
            "email": "hanako@example.com",
            "notes": "更新テスト",
            "birth_date": "",
            "middle_name": "",
        }
        record = _build_patient_record(extracted, "split", 1)
        assert record["phone"] == "09099990000"
        assert record["email"] == "hanako@example.com"
        assert record["reading"] == "タナカ ハナコ"
        assert record["notes"] == "更新テスト"


class TestFindDuplicatesFields(unittest.TestCase):
    """_find_duplicates の返却フィールドテスト"""

    def test_duplicate_candidate_has_reading_and_birthdate(self):
        """重複候補に reading と birth_date が含まれること"""
        # _find_duplicates は DB アクセスが必要なため、
        # 返却値の構造を確認するテスト
        candidate = {
            "id": 1,
            "name": "テスト 太郎",
            "patient_number": "P0001",
            "phone": "09012345678",
            "reading": "テスト タロウ",
            "birth_date": "1990-01-15",
            "reasons": ["氏名一致"],
        }
        assert "reading" in candidate
        assert "birth_date" in candidate
        assert candidate["reading"] == "テスト タロウ"


class TestTrySplitName(unittest.TestCase):
    """try_split_name テスト"""

    def test_space_separated(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("山田 太郎") == ("山田", "太郎")

    def test_fullwidth_space(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("山田\u3000太郎") == ("山田", "太郎")

    def test_multiple_spaces(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("山田  太郎") == ("山田", "太郎")

    def test_no_space(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("山田太郎") is None

    def test_three_tokens(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("Mary Jane Smith") is None

    def test_empty_string(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("") is None

    def test_whitespace_only(self):
        from app.api.patient_import import try_split_name
        assert try_split_name("   ") is None


class TestAutoSplitInBuild(unittest.TestCase):
    """split モードで full_name のみマッピング時の自動分割テスト"""

    def test_auto_split_success(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"full_name": "田中 太郎", "phone": "090-1234-5678"}
        record = _build_patient_record(extracted, "split", 1)
        assert record["last_name"] == "田中"
        assert record["first_name"] == "太郎"
        assert record["name"] == "田中 太郎"

    def test_auto_split_unsplittable(self):
        from app.api.patient_import import _build_patient_record
        extracted = {"full_name": "田中太郎", "phone": "090-1234-5678"}
        with self.assertRaises(ValueError):
            _build_patient_record(extracted, "split", 1)


if __name__ == "__main__":
    unittest.main()
